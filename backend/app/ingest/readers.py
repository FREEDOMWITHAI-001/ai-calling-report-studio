"""Tabular file readers.

Handles CSV (including embedded newlines inside quoted transcript fields) and
XLSX workbooks with many sheets, streaming rows so a 20 MB call log never has
to be held in memory twice.
"""
from __future__ import annotations

import csv
import sys
from pathlib import Path
from typing import Iterator

from openpyxl import load_workbook

csv.field_size_limit(min(sys.maxsize, 2**31 - 1))

CSV_EXT = {".csv", ".txt", ".tsv"}
XLSX_EXT = {".xlsx", ".xlsm"}


def is_excel(path: str | Path) -> bool:
    return Path(path).suffix.lower() in XLSX_EXT


def list_sheets(path: str | Path) -> list[dict]:
    """Sheet inventory for an uploaded workbook (or the single CSV 'sheet')."""
    path = Path(path)
    if not is_excel(path):
        return [{"name": path.name, "rows": None, "columns": None}]
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        return [
            {"name": ws.title, "rows": ws.max_row, "columns": ws.max_column}
            for ws in wb.worksheets
        ]
    finally:
        wb.close()


def _dedupe(headers: list[str]) -> list[str]:
    out: list[str] = []
    seen: dict[str, int] = {}
    for idx, raw in enumerate(headers):
        name = (str(raw).strip() if raw is not None else "") or f"column_{idx + 1}"
        name = " ".join(name.split())
        if name in seen:
            seen[name] += 1
            name = f"{name} ({seen[name]})"
        else:
            seen[name] = 0
        out.append(name)
    return out


def _sniff_delimiter(sample: str) -> str:
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
    except csv.Error:
        return ","


def _csv_rows(path: Path) -> Iterator[list]:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            with open(path, newline="", encoding=encoding) as fh:
                sample = fh.read(64 * 1024)
                fh.seek(0)
                delimiter = _sniff_delimiter(sample)
                for row in csv.reader(fh, delimiter=delimiter):
                    yield row
            return
        except UnicodeDecodeError:
            continue
    raise ValueError("Could not decode file with utf-8 or latin-1")


def _xlsx_rows(path: Path, sheet: str | None) -> Iterator[list]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.worksheets[0]
        for row in ws.iter_rows(values_only=True):
            yield list(row)
    finally:
        wb.close()


def _raw_rows(path: Path, sheet: str | None) -> Iterator[list]:
    if is_excel(path):
        yield from _xlsx_rows(path, sheet)
    else:
        yield from _csv_rows(path)


def _is_blank(row: list) -> bool:
    return all(cell is None or str(cell).strip() == "" for cell in row)


def read_table(path: str | Path, sheet: str | None = None, header_row: int | None = None):
    """Yield (headers, row_iterator).

    header_row is 1-based within the sheet; when omitted the first row holding
    at least two non-empty cells is used.
    """
    path = Path(path)
    rows = _raw_rows(path, sheet)
    headers: list[str] | None = None
    index = 0
    buffered: list[list] = []

    for row in rows:
        index += 1
        if headers is None:
            if header_row is not None:
                if index < header_row:
                    continue
                headers = _dedupe(row)
                continue
            non_empty = sum(1 for c in row if c is not None and str(c).strip() != "")
            if non_empty >= 2:
                headers = _dedupe(row)
            continue
        buffered.append(row)
        if len(buffered) >= 500:
            break

    if headers is None:
        return [], iter(())

    width = len(headers)

    def iterator():
        for buffered_row in buffered:
            if not _is_blank(buffered_row):
                yield _to_dict(headers, buffered_row, width)
        for tail_row in rows:
            if not _is_blank(tail_row):
                yield _to_dict(headers, tail_row, width)

    return headers, iterator()


def _to_dict(headers: list[str], row: list, width: int) -> dict:
    if len(row) < width:
        row = list(row) + [None] * (width - len(row))
    return {headers[i]: row[i] for i in range(width)}


def preview(path: str | Path, sheet: str | None = None, limit: int = 15) -> dict:
    headers, rows = read_table(path, sheet)
    sample = []
    for i, row in enumerate(rows):
        if i >= limit:
            break
        sample.append({k: _jsonable(v) for k, v in row.items()})
    return {"columns": headers, "rows": sample}


def _jsonable(value):
    if value is None:
        return None
    if isinstance(value, (int, float, str, bool)):
        return value
    return str(value)


def count_rows(path: str | Path, sheet: str | None = None) -> int:
    _, rows = read_table(path, sheet)
    return sum(1 for _ in rows)
