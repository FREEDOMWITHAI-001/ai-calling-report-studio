"""Generic Excel renderer.

It knows the five block kinds and nothing about any particular client. One sheet
per section, in template order. Adding a client format never touches this file.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

THIN = Side(style="thin")

# openpyxl number formats per block format code.
NUMFMT = {
    "int": "#,##0",
    "number": "#,##0.0",
    "money": '"₹"#,##0',
    "pct": "0.0%",
    "pp": '+0.0"pp";-0.0"pp"',
    "delta": "+0.0%;-0.0%",
    "multiple": '0.0"×"',
    "p": "0.0000",
    "text": "@",
}


def _hex(value: str) -> str:
    return (value or "").lstrip("#").upper() or "000000"


class Sheet:
    """Cursor over one worksheet, so blocks can just append."""

    def __init__(self, ws, brand: dict):
        self.ws = ws
        self.brand = brand
        self.row = 1
        self.widths: dict[int, int] = {}

    def _track(self, col: int, text: str) -> None:
        self.widths[col] = max(self.widths.get(col, 10), min(52, len(str(text)) + 3))

    def write(self, col: int, value, *, bold=False, size=None, color=None, fill=None,
              numfmt=None, align=None, border=False, wrap=False):
        cell = self.ws.cell(row=self.row, column=col, value=value)
        cell.font = Font(
            bold=bold,
            size=size or 10,
            color=_hex(color or self.brand.get("ink", "16141F")),
            name="Calibri",
        )
        if fill:
            cell.fill = PatternFill("solid", fgColor=_hex(fill))
        if numfmt:
            cell.number_format = numfmt
        cell.alignment = Alignment(
            horizontal=align or "left", vertical="center", wrap_text=wrap
        )
        if border:
            cell.border = Border(bottom=THIN)
        self._track(col, value if value is not None else "")
        return cell

    def blank(self, n: int = 1) -> None:
        self.row += n

    def finish(self) -> None:
        for col, width in self.widths.items():
            self.ws.column_dimensions[get_column_letter(col)].width = width
        self.ws.freeze_panes = "A2"


def _title(sheet: Sheet, text: str, subtitle: str | None = None) -> None:
    sheet.write(1, text, bold=True, size=14, color=sheet.brand.get("ink"))
    sheet.row += 1
    if subtitle:
        sheet.write(1, subtitle, size=9, color=sheet.brand.get("muted"), wrap=True)
        sheet.row += 1
    sheet.blank()


def _notes(sheet: Sheet, notes: list[str]) -> None:
    for note in notes or []:
        sheet.write(1, note, size=9, color=sheet.brand.get("muted"), wrap=True)
        sheet.row += 1


def _render_kpi(sheet: Sheet, block: dict) -> None:
    _title(sheet, block.get("title") or "", block.get("subtitle"))
    for item in block["items"]:
        sheet.write(1, item["label"], size=10, color=sheet.brand.get("muted"))
        tone = item.get("tone", "neutral")
        color = {
            "pos": sheet.brand.get("positive"),
            "crit": sheet.brand.get("negative"),
            "accent": sheet.brand.get("accent"),
        }.get(tone, sheet.brand.get("ink"))
        sheet.write(2, item["value"], bold=True, size=11, color=color,
                    numfmt=NUMFMT.get(item["fmt"], "General"), align="right")
        if item.get("note"):
            sheet.write(3, item["note"], size=9, color=sheet.brand.get("muted"))
        sheet.row += 1
    sheet.blank()
    _notes(sheet, block.get("notes"))
    sheet.blank()


def _render_funnel(sheet: Sheet, block: dict) -> None:
    _title(sheet, block.get("title") or "", block.get("subtitle"))
    for col, label in enumerate(("Stage", "Count", "Rate"), start=1):
        sheet.write(col, label, bold=True, size=9, color=sheet.brand.get("muted"),
                    fill=sheet.brand.get("band"), border=True,
                    align="right" if col > 1 else "left")
    sheet.row += 1
    for stage in block["stages"]:
        sheet.write(1, stage["label"], size=10,
                    color=sheet.brand.get("muted") if stage.get("dim") else sheet.brand.get("ink"))
        sheet.write(2, stage["value"], size=10, numfmt=NUMFMT["int"], align="right")
        sheet.write(3, stage.get("rate"), size=10, numfmt=NUMFMT["pct"], align="right")
        sheet.row += 1
    sheet.blank()
    _notes(sheet, block.get("notes"))
    sheet.blank()


def _render_table(sheet: Sheet, block: dict) -> None:
    _title(sheet, block.get("title") or "", block.get("subtitle"))
    columns = block["columns"]
    for index, column in enumerate(columns, start=1):
        sheet.write(index, column["label"], bold=True, size=9,
                    color=sheet.brand.get("muted"), fill=sheet.brand.get("band"),
                    border=True, align=column.get("align", "left"))
    sheet.row += 1

    emphasis = block.get("emphasis") or {}
    first_key = columns[0]["key"] if columns else None
    for row in block["rows"]:
        marker = emphasis.get(str(row.get(first_key)))
        fill = {"total": sheet.brand.get("band"),
                "baseline": sheet.brand.get("baseline"),
                "verdict": sheet.brand.get("band")}.get(marker)
        for index, column in enumerate(columns, start=1):
            key = column["key"]
            # A row may override the format of one cell (a delta row inside a
            # rate column, for instance).
            fmt = row.get(f"_fmt_{key}") or column.get("fmt", "text")
            sheet.write(
                index, row.get(key),
                bold=marker in ("total", "verdict"),
                size=10,
                fill=fill,
                numfmt=NUMFMT.get(fmt, "General") if fmt != "text" else None,
                align=column.get("align", "left"),
            )
        sheet.row += 1
    sheet.blank()
    _notes(sheet, block.get("notes"))
    sheet.blank()


def _render_text(sheet: Sheet, block: dict) -> None:
    if block.get("title"):
        _title(sheet, block["title"])
    sheet.write(1, block.get("body") or "", size=10, color=sheet.brand.get("muted"), wrap=True)
    sheet.row += 2


RENDERERS = {
    "kpi": _render_kpi,
    "funnel": _render_funnel,
    "table": _render_table,
    "matrix": _render_table,
    "text": _render_text,
}


def _sheet_name(title: str, used: set[str]) -> str:
    # Excel: 31 chars, no []:*?/\ , and unique.
    clean = "".join(ch for ch in title if ch not in "[]:*?/\\")[:31] or "Sheet"
    name, suffix = clean, 2
    while name.lower() in used:
        name = f"{clean[:28]}_{suffix}"
        suffix += 1
    used.add(name.lower())
    return name


def _render_cover(wb, meta: dict, doc: dict, brand: dict, used: set) -> None:
    """The title sheet: what this report is, and the headline counts."""
    cover = Sheet(wb.create_sheet(_sheet_name("Cover", used)), brand)
    cover.write(1, meta.get("cover_title") or meta["title"], bold=True, size=18,
                color=brand.get("accent"))
    cover.row += 2
    for label, value in [
        ("Client", meta["client"]),
        ("Format", meta["template_label"]),
        ("Window", f"{meta['date_from']} to {meta['date_to']} ({meta['window_days']} days)"),
        ("Language", meta.get("language") or "all"),
        ("Generated", meta["generated_at"]),
    ]:
        cover.write(1, label, size=10, color=brand.get("muted"))
        cover.write(2, value, size=10, bold=True)
        cover.row += 1
    cover.blank()
    totals = doc.get("totals") or {}
    cover.write(1, "At a glance", bold=True, size=12)
    cover.row += 1
    for label, key, fmt in [
        ("Registrants", "registrants", "int"),
        ("People dialled", "dialled", "int"),
        ("People reached", "reached", "int"),
        ("Attendees", "attended", "int"),
        ("Buyers", "buyers", "int"),
        ("Calls placed", "calls_placed", "int"),
        ("Talk cost", "talk_cost", "money"),
        ("Revenue recorded", "revenue", "money"),
    ]:
        cover.write(1, label, size=10, color=brand.get("muted"))
        cover.write(2, totals.get(key), size=10, bold=True, numfmt=NUMFMT[fmt], align="right")
        cover.row += 1

    if doc.get("skipped"):
        cover.blank()
        cover.write(1, "Sections not included", bold=True, size=11,
                    color=brand.get("negative"))
        cover.row += 1
        for item in doc["skipped"]:
            cover.write(1, item["title"], size=10)
            cover.write(2, item["reason"], size=9, color=brand.get("muted"))
            cover.row += 1
    cover.finish()


def build_workbook(doc: dict, path: str | Path) -> Path:
    """Render a composed document to XLSX. Returns the path written."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    brand = doc.get("brand") or {}
    wb = Workbook()
    wb.remove(wb.active)
    used: set[str] = set()

    meta = doc["meta"]
    # A format can opt out of the cover when its layout is a fixed set of sheets.
    if meta.get("cover", True):
        _render_cover(wb, meta, doc, brand, used)

    # One sheet per available section
    for section in doc["sections"]:
        if not section["available"]:
            continue
        sheet = Sheet(wb.create_sheet(_sheet_name(section["title"], used)), brand)
        for block in section["blocks"]:
            RENDERERS.get(block["kind"], _render_text)(sheet, block)
        sheet.finish()

    if brand.get("footer"):
        for ws in wb.worksheets:
            ws.oddFooter.left.text = brand["footer"]

    wb.save(path)
    return path
