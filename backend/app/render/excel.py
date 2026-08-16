"""Excel output: 'Overview' summary tab + a detail tab that mirrors the
'CBA X report' tab of the reference workbooks, plus a raw-numbers tab."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import style as S

MONEY_FMT = '"₹"#,##0'
PCT_FMT = "0.0%"
PCT2_FMT = "0.00%"

THIN = Side(style="thin", color=S.RULE)


def _title(ws, row: int, text: str, size: int = 15) -> int:
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(bold=True, size=size, color=S.INK)
    return row + 1


def _subtitle(ws, row: int, text: str) -> int:
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(size=10, color=S.MUTED, italic=True)
    return row + 1


def _section(ws, row: int, text: str, width: int = 8) -> int:
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(bold=True, size=11, color="FFFFFF")
    for col in range(1, width + 1):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=S.INK)
    return row + 1


def _header_row(ws, row: int, headers: list[str]) -> int:
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=idx, value=header)
        cell.font = Font(bold=True, size=10, color=S.INK)
        cell.fill = PatternFill("solid", fgColor=S.ACCENT_SOFT)
        cell.border = Border(bottom=THIN)
        cell.alignment = Alignment(horizontal="right" if idx > 1 else "left")
    return row + 1


def _note(ws, row: int, text: str, width: int = 8, height: int = 60) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=width)
    cell = ws.cell(row=row, column=1, value=text)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    cell.font = Font(size=9, color=S.MUTED)
    ws.row_dimensions[row].height = height
    return row + 2


def _group_table(ws, row: int, rows: list[dict], baseline_key: str = "baseline") -> int:
    row = _header_row(ws, row, [
        "Group", "Registrants", "Showed", "Show-up %", "Show-up Δ", "Buyers", "Buyer %", "Buyer Δ",
    ])
    for item in rows:
        is_total = item.get("key") == "total" or item.get("label") == "Total"
        is_baseline = item.get("key") == baseline_key
        values = [
            item["label"],
            item["registrants"],
            item["showed"],
            item["show_rate"],
            item["show_delta"],
            item["buyers"],
            item["buy_rate"],
            item["buy_delta"],
        ]
        for idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=idx, value="—" if value is None and idx in (5, 8) else value)
            cell.font = Font(bold=is_total, size=10)
            if is_baseline:
                cell.fill = PatternFill("solid", fgColor=S.BASELINE_BG)
            elif is_total:
                cell.fill = PatternFill("solid", fgColor=S.BAND)
            if idx in (4, 7):
                cell.number_format = PCT2_FMT if idx == 7 else PCT_FMT
            elif idx in (5, 8) and value is not None:
                cell.number_format = "+0.0%;-0.0%"
                cell.font = Font(size=10, color=S.POSITIVE if value >= 0 else S.NEGATIVE)
            elif idx in (2, 3, 6):
                cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="left" if idx == 1 else "right")
        row += 1
    return row + 1


def _kv(ws, row: int, label: str, value, fmt: str | None = None, bold: bool = False) -> int:
    cell = ws.cell(row=row, column=1, value=label)
    cell.font = Font(bold=bold, size=10)
    out = ws.cell(row=row, column=4, value=value)
    out.font = Font(bold=True, size=11, color=S.INK)
    out.alignment = Alignment(horizontal="right")
    if fmt:
        out.number_format = fmt
    return row + 1


def build_excel(result: dict, path: str | Path) -> Path:
    path = Path(path)
    head = result["headline"]
    meta = result["meta"]
    groups = result["groups"]

    wb = Workbook()

    # ---------------- Overview ---------------- #
    ws = wb.active
    ws.title = "Overview"
    ws.sheet_view.showGridLines = False
    for col, width in zip("ABCDEF", (46, 20, 18, 16, 16, 12)):
        ws.column_dimensions[col].width = width

    row = _title(ws, 1, f"COACHEASILY — WHAT AI CALLING ADDED  ({meta['title']})", 16)
    row = _subtitle(ws, row, S.subtitle_line(result))
    row += 1
    row = _section(ws, row, "REVENUE WITH AI CALLING vs WITHOUT", 6)
    row = _header_row(ws, row, [
        "Program", "Revenue without AI", "Revenue with AI", "AI added", "Relative uplift", "ROI",
    ])
    program_label = meta["title"]
    for label, bold in ((program_label, False), ("Total", True)):
        ws.cell(row=row, column=1, value=label).font = Font(bold=bold, size=10)
        for idx, value, fmt in (
            (2, head["revenue_without_ai"], MONEY_FMT),
            (3, head["revenue_with_ai"], MONEY_FMT),
            (4, head["revenue_added"], MONEY_FMT),
            (5, head["relative_uplift"], PCT_FMT),
        ):
            cell = ws.cell(row=row, column=idx, value=value)
            cell.number_format = fmt
            cell.font = Font(bold=bold, size=10)
            cell.alignment = Alignment(horizontal="right")
        roi = ws.cell(row=row, column=6, value=S.multiple(head["roi"]))
        roi.font = Font(bold=True, size=10, color=S.ACCENT)
        roi.alignment = Alignment(horizontal="right")
        if bold:
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=S.BAND)
        row += 1

    row += 1
    row = _note(ws, row, S.method_note(result), width=6, height=78)

    row = _section(ws, row, "EXTRA SALES — WEIGHTED BY LEAD AGE (like-for-like)", 6)
    row = _header_row(ws, row, [
        "Lead age band", "Connected", "Connected buy %", "Baseline", "Baseline buy %", "Extra sales",
    ])
    for band in result["bands"]:
        values = [
            band["band"], band["connected"], band["connected_buy_rate"],
            band["baseline"], band["baseline_buy_rate"], round(band["extra_sales"], 1),
        ]
        for idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=idx, value=value)
            cell.font = Font(size=10)
            if idx in (3, 5):
                cell.number_format = PCT2_FMT
            elif idx in (2, 4):
                cell.number_format = "#,##0"
            elif idx == 6:
                cell.number_format = "0.0"
            cell.alignment = Alignment(horizontal="left" if idx == 1 else "right")
        row += 1
    ws.cell(row=row, column=1, value="Total extra sales credited to AI").font = Font(bold=True, size=10)
    total_cell = ws.cell(row=row, column=6, value=round(head["extra_sales"], 1))
    total_cell.font = Font(bold=True, size=10)
    total_cell.number_format = "0.0"
    total_cell.alignment = Alignment(horizontal="right")

    # ---------------- Detail ---------------- #
    detail = wb.create_sheet(meta["title"][:28] or "Detail")
    detail.sheet_view.showGridLines = False
    for col, width in zip("ABCDEFGH", (40, 14, 12, 12, 12, 12, 12, 12)):
        detail.column_dimensions[col].width = width

    row = _title(detail, 1, f"COACHEASILY — {meta['title']}  ·  AI calling impact", 15)
    row = _subtitle(detail, row, S.subtitle_line(result))
    row += 1
    row = _section(detail, row, "BUSINESS IMPACT — with AI vs without")
    row = _kv(detail, row, "Revenue without AI calling", head["revenue_without_ai"], MONEY_FMT)
    row = _kv(detail, row, "Revenue with AI calling", head["revenue_with_ai"], MONEY_FMT)
    row = _kv(detail, row, "AI calling added", head["revenue_added"], MONEY_FMT, bold=True)
    row = _kv(detail, row, "Relative revenue increase", head["relative_uplift"], PCT_FMT)
    row = _kv(detail, row, "Extra sales credited to AI (weighted, like-for-like)",
              round(head["extra_sales"], 1), "0.0")
    row = _kv(detail, row, "Sale value", head["sale_value"], MONEY_FMT)
    row = _kv(detail, row,
              f"Talk-minutes × {S.money(head['params']['cost_per_minute'] if 'params' in head else meta['params']['cost_per_minute'], 2)}",
              head["talk_cost"], MONEY_FMT)
    row = _kv(detail, row, "ROI (return multiple)", S.multiple(head["roi"]), bold=True)
    row += 1

    row = _section(detail, row, "SHOW-UP & BUYERS — by bot reached (whole window · Δ = vs baseline)")
    ordered = [
        {**groups["total"], "key": "total"},
        {**groups["signup"], "key": "signup"},
        {**groups["day_of"], "key": "day_of"},
        {**groups["both"], "key": "both"},
        {**groups["baseline"], "key": "baseline"},
    ]
    row = _group_table(detail, row, ordered)
    row = _note(detail, row, S.narrative(result), height=88)

    row = _section(detail, row, "PER-CALL-DAY DETAIL  (each webinar day, by bot reached · Δ = vs that day's baseline)")
    row += 1
    for day in result["daily"]:
        cell = detail.cell(row=row, column=1, value=S.pretty_date(day["date"]))
        cell.font = Font(bold=True, size=11, color=S.ACCENT)
        row += 1
        row = _group_table(detail, row, day["rows"])

    row = _section(detail, row, "METHOD, SOURCES & DATA AUDIT")
    audit = result["audit"]
    calls = result["calls"]
    sig = result["significance"]
    entries = [
        ("Window", f"{S.pretty_date(meta['date_from'])} – {S.pretty_date(meta['date_to'])} "
                   f"({meta['window_days']} days) by REGISTRATION date."),
        ("Registrants", f"{audit['registration_rows']:,} rows in the window → "
                        f"{audit['team_registration_rows']:,} team/test rows and "
                        f"{audit['repeat_registration_rows']:,} repeat registrations removed → "
                        f"{audit['unique_registrants']:,} unique people, matched across files by "
                        f"phone OR email OR name."),
        ("Connected", f"Duration MORE THAN {audit['connected_threshold_s']}s. Of {calls['calls_placed']:,} calls "
                      f"the in-scope bots placed in the window, {calls['calls_connected']:,} connected, reaching "
                      f"{result['headline']['connected_people']:,} distinct registrants. "
                      f"{calls['calls_never_connected']:,} calls never connected at all. "
                      f"{calls['matched_calls']:,} of {calls['calls_placed']:,} call rows "
                      f"({S.pct(calls['match_rate'], 0)}) matched a window registrant."),
        ("Buyers", f"Every qualifying sale row is counted as one FULL sale of {S.money(head['sale_value'])}; a person "
                   f"who locked and then paid the balance is counted ONCE. "
                   f"{audit['sale_rows_in_window']:,} sale rows in window · "
                   f"{audit['sale_rows_outside_cohort']:,} belong to people outside this cohort · "
                   f"{audit['sale_rows_before_registration']:,} dated before the person registered were dropped."),
        ("Show-up", f"Per-person match against the attendance data → {head['showed']:,}/{head['registrants']:,} = "
                    f"{S.pct(head['showed'] / head['registrants'] if head['registrants'] else 0, 0)}."
                    + (f" Platform's own count for these days is {audit['platform_leads']:,} leads / "
                       f"{audit['platform_show_up']:,} show-ups." if audit.get("platform_leads") else "")),
        ("Cost", f"{calls['calls_with_audio']:,} calls had someone on the line, holding "
                 f"{calls['talk_seconds']:,} seconds = {calls['talk_minutes_exact']:,} minutes of real talk time. "
                 f"Billing is PER MINUTE, so each call is rounded UP: {calls['billed_minutes']:,} billed minutes × "
                 f"{S.money(calls['cost_per_minute'], 2)} = {S.money(calls['talk_cost'])}."),
        ("Extra sales — the weighted calculation",
         f"Registrants are split into {len(result['bands'])} bands by how many days they had before the data was "
         f"pulled. Inside each band, connected leads are compared only with baseline leads of the SAME age, and the "
         f"band's gap is multiplied by that band's connected count "
         f"({' + '.join(f'{b['extra_sales']:.1f}' for b in result['bands'])} = {head['extra_sales']:.1f} sales). "
         f"Weighted gap {head['weighted_gap_points'] * 100:.2f} points vs "
         f"{head['simple_gap_points'] * 100:.2f} points unweighted."),
        ("Significance", f"Show-up {S.p_value(sig['show_up']['p_value'])} · "
                         f"Buying {S.p_value(sig['buying']['p_value'])} (two-proportion z-test, "
                         f"connected vs baseline)."),
        ("THE BIG CAVEAT",
         "This is an OBSERVATIONAL comparison, not an experiment. Nobody was randomly left uncalled, so 'connected' "
         "people are self-selected — they answered their phone, which already marks them as more engaged. Part of "
         "their higher show-up and buy rate would have happened anyway. Normalising for lead age removes the one "
         "bias that IS measurable; this one cannot be removed without a hold-out test."),
    ]
    for label, text in entries:
        cell = detail.cell(row=row, column=1, value=label)
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        detail.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        body = detail.cell(row=row, column=2, value=text)
        body.alignment = Alignment(wrap_text=True, vertical="top")
        body.font = Font(size=9, color=S.MUTED)
        detail.row_dimensions[row].height = max(30, 12 * (len(text) // 95 + 1))
        row += 1

    # ---------------- Machine-readable numbers ---------------- #
    raw = wb.create_sheet("Numbers")
    raw.append(["metric", "value"])
    raw.cell(row=1, column=1).font = Font(bold=True)
    raw.cell(row=1, column=2).font = Font(bold=True)
    flat = {
        **{f"headline.{k}": v for k, v in head.items() if not isinstance(v, dict)},
        **{f"calls.{k}": v for k, v in calls.items() if not isinstance(v, dict)},
        **{f"audit.{k}": v for k, v in audit.items()},
    }
    for key, value in flat.items():
        raw.append([key, value])
    raw.column_dimensions["A"].width = 40
    raw.column_dimensions["B"].width = 24

    bots = wb.create_sheet("Bot cost")
    bots.append(["Bot", "Role", "Calls with talk time", "Billed minutes", "Cost", "Days active"])
    for col in range(1, 7):
        bots.cell(row=1, column=col).font = Font(bold=True)
    for name, bucket in result["calls"]["by_bot"].items():
        bots.append([name, bucket.get("role"), bucket.get("calls"), bucket.get("billed_minutes"),
                     bucket.get("cost"), bucket.get("days_active")])
    for idx, width in enumerate((42, 12, 20, 16, 14, 14), start=1):
        bots.column_dimensions[get_column_letter(idx)].width = width

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path
